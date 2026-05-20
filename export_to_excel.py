import pymysql
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'Ida2451212073',
    'database': 'movies',
    'charset': 'utf8mb4'
}
OUTPUT_PATH = r"douban_comments_export.xlsx"


def _sanitize(value):
    if value is None:
        return ''
    return str(value)


def export():
    conn = pymysql.connect(**DB_CONFIG)

    sql = """
        SELECT
            m.title_cn,
            c.content,
            c.time,
            c.star
        FROM comments_list c
        JOIN movies_list m ON c.movie_id = m.id
        ORDER BY m.title_cn, c.time
    """
    df = pd.read_sql(sql, conn)
    conn.close()

    if df.empty:
        return

    movies = df['title_cn'].dropna().unique()

    with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
        for movie in movies:
            sheet_df = df[df['title_cn'] == movie][
                ['title_cn', 'content', 'time', 'star']
            ].copy()
            sheet_df.columns = ['movie_name', 'comment', 'comment_time', 'star']
            sheet_df.to_excel(writer, index=False, sheet_name=_sheet_name(movie))

    print(f"Export completed. Total {len(movies)} movies saved to: {OUTPUT_PATH}")

    _format(OUTPUT_PATH, len(movies))


def _sheet_name(name):
    name = str(name).strip()[:31]
    for c in r'\/:*?[]':
        name = name.replace(c, '')
    return name or 'Sheet1'


def _format(path, num_sheets):
    wb = load_workbook(path)

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10

        ws.row_dimensions[1].height = 28

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            row[3].alignment = Alignment(horizontal='center', vertical='top')

        light_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        for row in ws.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = light_fill

    wb.save(path)


if __name__ == '__main__':
    export()
